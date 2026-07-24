"""Tests for make_report. Run with:  python -m pytest"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from make_report import build_report, read_data  # noqa: E402


def _sample_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "region": ["West", "East", "North"],
            "product": ["A", "B", "C"],
            "units": [10, -2, 5],
            "amount": [2500, -440, 900],
        }
    )


def test_build_report_creates_file(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    build_report(_sample_df(), out, chart_col="amount", title="Test")
    assert out.exists()


def test_report_has_title_header_and_totals(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    build_report(_sample_df(), out, title="Q1 Sales")
    wb = load_workbook(out)
    ws = wb.active

    # Title in row 1
    assert ws.cell(row=1, column=1).value == "Q1 Sales"
    # Header in row 3
    assert ws.cell(row=3, column=1).value == "region"
    assert ws.cell(row=3, column=4).value == "amount"
    # Totals row after 3 data rows (rows 4-6) -> row 7
    assert ws.cell(row=7, column=1).value == "TOTAL"
    # Numeric total columns use SUM formulas
    assert str(ws.cell(row=7, column=3).value).startswith("=SUM(")
    assert str(ws.cell(row=7, column=4).value).startswith("=SUM(")


def test_header_is_frozen(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    build_report(_sample_df(), out)
    ws = load_workbook(out).active
    # Freeze at row 4 means rows 1-3 stay visible
    assert ws.freeze_panes == "A4"


def test_empty_dataframe_raises(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    try:
        build_report(pd.DataFrame(), out)
    except ValueError:
        return
    raise AssertionError("Expected ValueError for empty input")


def test_read_csv(tmp_path: Path) -> None:
    csv = tmp_path / "d.csv"
    csv.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
    df = read_data(csv, sheet=None)
    assert list(df.columns) == ["a", "b"]
    assert len(df) == 2
