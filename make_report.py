"""
make_report.py — Turn a raw CSV/Excel export into a formatted Excel report.

Takes a plain data file and produces a polished, presentation-ready workbook:
a styled header, auto-sized columns, a frozen header row, a totals row with live
SUM formulas for numeric columns, conditional formatting (negatives in red, high
values in green), and an optional bar chart — all in one command.

Usage:
    python make_report.py <input> [-o report.xlsx] [--chart COLUMN] [--title TEXT]

Examples:
    python make_report.py sales_raw.csv
    python make_report.py sales_raw.csv -o march_report.xlsx --chart amount --title "March Sales"
    python make_report.py data.xlsx --highlight-min 1000

Author: Synth88Labs
License: MIT
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- Theme (ExcelGuru green) ---
HEADER_FILL = PatternFill("solid", fgColor="217346")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, size=16, color="217346")
TOTAL_FILL = PatternFill("solid", fgColor="E8F3EC")
TOTAL_FONT = Font(bold=True, color="16281F")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="0E6245")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
_THIN = Side(style="thin", color="D9E2DC")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEADER_ROW = 3  # row 1 = title, row 2 = spacer, row 3 = column headers


def read_data(path: Path, sheet: str | None) -> pd.DataFrame:
    """Read a CSV or Excel file into a DataFrame, inferring numeric types."""
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    sheet_arg = sheet if sheet is not None else 0
    return pd.read_excel(path, sheet_name=sheet_arg)


def _write_title(ws: Worksheet, title: str, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left")


def _write_header(ws: Worksheet, columns: list[str]) -> None:
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=HEADER_ROW, column=c, value=str(name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _write_data(ws: Worksheet, df: pd.DataFrame) -> int:
    """Write the data rows; return the last data row index."""
    for r, (_, row) in enumerate(df.iterrows(), start=HEADER_ROW + 1):
        for c, name in enumerate(df.columns, start=1):
            value = row[name]
            # pandas NaN -> blank cell
            cell = ws.cell(row=r, column=c, value=None if pd.isna(value) else value)
            cell.border = BORDER
    return HEADER_ROW + len(df)


def _write_totals(ws: Worksheet, df: pd.DataFrame, last_row: int, numeric_cols: list[int]) -> None:
    total_row = last_row + 1
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER
        if c == 1:
            cell.value = "TOTAL"
        elif c in numeric_cols:
            col = get_column_letter(c)
            cell.value = f"=SUM({col}{HEADER_ROW + 1}:{col}{last_row})"


def _apply_conditional_formatting(
    ws: Worksheet, last_row: int, numeric_cols: list[int], highlight_min: float
) -> None:
    for c in numeric_cols:
        col = get_column_letter(c)
        rng = f"{col}{HEADER_ROW + 1}:{col}{last_row}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT)
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(highlight_min)],
                fill=GREEN_FILL,
                font=GREEN_FONT,
            ),
        )


def _autosize_and_freeze(ws: Worksheet, df: pd.DataFrame) -> None:
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)
    for c, name in enumerate(df.columns, start=1):
        values = [str(name)] + [str(v) for v in df[name].tolist()]
        width = max(len(v) for v in values) + 4
        ws.column_dimensions[get_column_letter(c)].width = min(width, 40)


def _add_chart(ws: Worksheet, df: pd.DataFrame, chart_col: str, last_row: int) -> None:
    idx = list(df.columns).index(chart_col) + 1
    chart = BarChart()
    chart.title = f"{chart_col} by row"
    chart.height, chart.width = 8, 16
    data = Reference(ws, min_col=idx, min_row=HEADER_ROW, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=HEADER_ROW + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{get_column_letter(len(df.columns) + 2)}{HEADER_ROW}")


def build_report(
    df: pd.DataFrame,
    out_path: Path,
    chart_col: str | None = None,
    title: str = "Report",
    highlight_min: float = 2000,
) -> None:
    """Build a formatted Excel report from *df* and save it to *out_path*."""
    if df.empty:
        raise ValueError("Input has no rows — nothing to report.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    numeric_cols = [
        i for i, n in enumerate(df.columns, start=1) if pd.api.types.is_numeric_dtype(df[n])
    ]

    _write_title(ws, title, len(df.columns))
    _write_header(ws, list(df.columns))
    last_row = _write_data(ws, df)
    _write_totals(ws, df, last_row, numeric_cols)
    _apply_conditional_formatting(ws, last_row, numeric_cols, highlight_min)
    _autosize_and_freeze(ws, df)

    if chart_col:
        if chart_col not in df.columns:
            print(f"  ! Chart column '{chart_col}' not found — skipping chart.", file=sys.stderr)
        else:
            _add_chart(ws, df, chart_col, last_row)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Turn a raw CSV/Excel export into a formatted Excel report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("input", type=Path, help="Input .csv or .xlsx file.")
    p.add_argument(
        "-o", "--output", type=Path, default=Path("report.xlsx"),
        help="Output .xlsx path. Default: report.xlsx",
    )
    p.add_argument("--chart", default=None, help="Numeric column to chart as a bar chart.")
    p.add_argument("--title", default="Report", help="Title shown at the top of the report.")
    p.add_argument(
        "--highlight-min", type=float, default=2000,
        help="Highlight numeric values >= this threshold in green. Default: 2000",
    )
    p.add_argument("--sheet", default=None, help="For Excel input: sheet name (default: first).")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if not args.input.is_file():
        print(f"Error: '{args.input}' is not a file.", file=sys.stderr)
        return 1

    try:
        df = read_data(args.input, args.sheet)
    except Exception as exc:  # noqa: BLE001
        print(f"Error reading '{args.input}': {exc}", file=sys.stderr)
        return 1

    try:
        build_report(df, args.output, args.chart, args.title, args.highlight_min)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(
        f"Report saved: {args.output.resolve()}  "
        f"({len(df)} rows, {len(df.columns)} columns)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
